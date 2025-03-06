import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font

def format_excel_from_csv(csv_file, final_excel_file):
    """
    Reads a CSV file, converts it to an Excel file, applies conditional formatting
    and header styling, then saves the final output as final_excel_file.
    
    Conditional formatting:
      - Applies classic red-to-green gradients (with a white midpoint) starting from row 2.
      - Uses red-to-green formatting on columns D, H, and K, and the reverse for column Q.
      
    Header styling:
      - Rotates the header row 45° clockwise.
      - Unbolds the entire header row, then bolds header cells in columns D, H, K, and Q.
      - Adjusts column A width to fit its content.
      - Sets other columns (except A) to a fixed width of 7.
    """
    temp_excel_file = "temp_summary.xlsx"  # Temporary Excel file created from CSV

    # Read CSV data and write to a temporary Excel file
    df = pd.read_csv(csv_file)
    df.to_excel(temp_excel_file, index=False)  # Save without the DataFrame index

    # Remove existing final file if it exists
    if os.path.exists(final_excel_file):
        os.remove(final_excel_file)
        print(f"Existing file deleted: {final_excel_file}")

    # Load the Excel file created from CSV (temporary file)
    wb = load_workbook(temp_excel_file)
    ws = wb.active  # Use the active sheet

    # Define columns for the two color scales
    columns_red_to_green = ["D", "H", "K"]  # Low = Red, High = Green
    columns_green_to_red = ["Q"]            # Low = Green, High = Red

    # Color scale 1: Classic red-to-green with white midpoint
    color_scale_red_to_green = ColorScaleRule(
        start_type="min", start_color="F8696B",    # Red for minimum
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",  # White for median
        end_type="max", end_color="63BE7B"           # Green for maximum
    )

    # Color scale 2: Reverse scale (Green to Red) with white midpoint
    color_scale_green_to_red = ColorScaleRule(
        start_type="min", start_color="63BE7B",    # Green for minimum
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",  # White for median
        end_type="max", end_color="F8696B"           # Red for maximum
    )

    # Function to apply conditional formatting to the given columns, starting from row 2
    def apply_conditional_formatting(columns, color_scale):
        for col in columns:
            col_range = f"{col}2:{col}{ws.max_row}"  # Apply from row 2 to the last used row
            ws.conditional_formatting.add(col_range, color_scale)

    # Apply conditional formatting
    apply_conditional_formatting(columns_red_to_green, color_scale_red_to_green)
    apply_conditional_formatting(columns_green_to_red, color_scale_green_to_red)

    # Rotate header row (row 1) 45 degrees clockwise
    for cell in ws[1]:
        cell.alignment = Alignment(textRotation=45)

    # Unbold the entire header row first
    for cell in ws[1]:
        cell.font = Font(bold=False)

    # Bold specific header cells in columns D, H, K, and Q
    for col in ["D", "H", "K", "Q"]:
        ws[f"{col}1"].font = Font(bold=True)

    # Adjust column A width to fit its content
    max_length = 0
    for cell in ws["A"]:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))
    # Optionally, add a little extra space (e.g., 2 units)
    ws.column_dimensions["A"].width = max_length + 2

    # Set widths for the remaining columns (skip column A)
    for col in ws.columns:
        if col[0].column_letter != "A":
            ws.column_dimensions[col[0].column_letter].width = 7

    # Save as the final file (overwriting if necessary)
    wb.save(final_excel_file)
    print(f"Conditional formatting applied successfully! New file saved as: {final_excel_file}")

    # Optionally, remove the temporary Excel file
    os.remove(temp_excel_file)

    # Open the final file (works on Windows)
    os.startfile(final_excel_file)
