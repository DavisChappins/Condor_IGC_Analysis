import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule

# Load the CSV file
df = pd.read_csv('summary.csv')

# Save the DataFrame to an Excel file
df.to_excel('summary.xlsx', index=False)

# Load the workbook
wb = load_workbook('summary.xlsx')

# Select the active worksheet
ws = wb.active

# Define the font style for the header row
header_font = Font(name='Calibri', size=11, bold=True, color='000000', italic=False, vertAlign=None, underline='none', strike=False)

# Define the border style
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

# Define the alignment style
alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)

# Apply the font style, border, and alignment to the header row
for cell in ws[1]:
    cell.font = header_font
    cell.border = border
    cell.alignment = alignment
    
# Extract the values from column G as a list of integers
values = [int(cell.value) for cell in ws['G'][2:]]

# Find the smallest and largest values in the list
min_value = min(values)
max_value = max(values)


# Calculate the middle value
mid_value = (min_value + max_value) / 2

# Define the color scale rule
color_scale_rule = ColorScaleRule(start_type='num', start_value=min_value, start_color='FF0000', 
                                  mid_type='num', mid_value=4, mid_color='FFFFFF',
                                  end_type='num', end_value=max_value, end_color='00FF00')

# Apply the color scale rule to the range G2 to the end of data in column G
ws.conditional_formatting.add('G2:G' + str(ws.max_row), color_scale_rule)

# Save the workbook
wb.save('summary.xlsx')